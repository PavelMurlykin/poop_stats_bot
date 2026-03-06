"""Сервис формирования XLSX-отчёта по статистике пользователя."""

import io
from datetime import date, datetime
from typing import Any, TypeAlias

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from config import DATE_FORMAT_DISPLAY, DATE_FORMAT_STORAGE
from db.repositories import fetch_all_for_report

RowData: TypeAlias = dict[str, Any]
RowsData: TypeAlias = list[RowData]

DATE_COLUMN_INDEXES = {1, 12, 13}
NUMBER_COLUMN_INDEXES = {5, 7, 9, 15}
CENTERED_COLUMN_INDEXES = DATE_COLUMN_INDEXES | NUMBER_COLUMN_INDEXES

DATE_COLUMN_WIDTH = 12
COUNT_COLUMN_WIDTH = 12
SHORT_TEXT_COLUMN_WIDTH = 18
LONG_TEXT_COLUMN_WIDTH = 24
XL_TEXT_COLUMN_WIDTH = 28

COLUMN_WIDTHS = {
    1: DATE_COLUMN_WIDTH,
    2: SHORT_TEXT_COLUMN_WIDTH,
    3: SHORT_TEXT_COLUMN_WIDTH,
    4: SHORT_TEXT_COLUMN_WIDTH,
    5: COUNT_COLUMN_WIDTH,
    6: SHORT_TEXT_COLUMN_WIDTH,
    7: COUNT_COLUMN_WIDTH,
    8: LONG_TEXT_COLUMN_WIDTH,
    9: COUNT_COLUMN_WIDTH,
    10: XL_TEXT_COLUMN_WIDTH,
    11: LONG_TEXT_COLUMN_WIDTH,
    12: DATE_COLUMN_WIDTH,
    13: DATE_COLUMN_WIDTH,
    14: LONG_TEXT_COLUMN_WIDTH,
    15: COUNT_COLUMN_WIDTH,
}

BRISTOL = {
    0: 'Отсутствие дефекации',
    1: 'Отдельные твёрдые комки, как орехи, трудно проходят [запор]',
    2: 'Колбасовидный, но комковатый [запор или склонность к запору]',
    3: 'Колбасовидный с трещинами на поверхности [норма]',
    4: 'Колбасовидный гладкий и мягкий [норма]',
    5: 'Мягкие маленькие шарики с чёткими краями [склонность к диарее]',
    6: 'Рыхлые кусочки с неровными краями, кашицеобразный [диарея]',
    7: 'Водянистый, без твёрдых кусочков [сильная диарея]',
}

REPORT_COLUMNS = [
    'Дата',
    'Завтрак',
    'Обед',
    'Ужин',
    'Количество перекусов',
    'Перекусы',
    'Количество приемов лекарств',
    'Лекарства',
    'Количество походов в туалет',
    'Качество стула',
    'Самочувствие',
    'Подъем',
    'Отход ко сну',
    'Качество сна',
    'Стаканов воды',
]


def _to_display(date_value: date | datetime | str) -> str:
    """Преобразует дату из формата хранения в `ДД.ММ.ГГГГ`.

    Args:
        date_value: Дата в виде `date`, `datetime` или строки `YYYY-MM-DD`.

    Returns:
        str: Дата в пользовательском формате отображения.
    """
    if isinstance(date_value, datetime):
        return date_value.strftime(DATE_FORMAT_DISPLAY)
    if isinstance(date_value, date):
        return date_value.strftime(DATE_FORMAT_DISPLAY)
    parsed_date = datetime.strptime(str(date_value), DATE_FORMAT_STORAGE)
    return parsed_date.strftime(DATE_FORMAT_DISPLAY)


def _apply_worksheet_style(
    worksheet,
    total_rows: int,
    total_columns: int,
) -> None:
    """Применяет оформление к листу Excel-отчёта.

    Args:
        worksheet: Рабочий лист openpyxl.
        total_rows: Количество строк таблицы вместе с заголовком.
        total_columns: Количество колонок в таблице.
    """
    border_side = Side(style='thin', color='000000')
    border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )
    header_font = Font(bold=True)
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )
    text_alignment = Alignment(
        horizontal='left',
        vertical='top',
        wrap_text=True,
    )
    centered_alignment = Alignment(horizontal='center', vertical='center')

    for column_index, width in COLUMN_WIDTHS.items():
        column_name = get_column_letter(column_index)
        worksheet.column_dimensions[column_name].width = width

    worksheet.row_dimensions[1].height = 36
    worksheet.freeze_panes = 'A2'

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=total_rows,
        min_col=1,
        max_col=total_columns,
    ):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = header_font
                cell.alignment = header_alignment
            elif cell.column in CENTERED_COLUMN_INDEXES:
                cell.alignment = centered_alignment
            else:
                cell.alignment = text_alignment


def _group_rows_by_date(rows: RowsData) -> dict[Any, RowsData]:
    """Группирует список словарей по полю `date`.

    Args:
        rows: Список строк, содержащих ключ `date`.

    Returns:
        dict[Any, RowsData]: Словарь `дата -> список строк`.
    """
    grouped_rows: dict[Any, RowsData] = {}
    for row in rows:
        grouped_rows.setdefault(row['date'], []).append(row)
    return grouped_rows


def _extract_meal_columns(
    meals_for_day: RowsData,
) -> tuple[str, str, str, int, str]:
    """Формирует колонки отчёта по питанию для одной даты.

    Args:
        meals_for_day: Список приемов пищи за день.

    Returns:
        tuple[str, str, str, int, str]: Завтрак, обед, ужин, количество
        перекусов и строка с перекусами.
    """
    breakfast = ''
    lunch = ''
    dinner = ''
    snacks: list[str] = []

    for meal in meals_for_day:
        meal_type = meal['meal_type']
        description = meal['description']
        if meal_type == 'breakfast':
            breakfast = description
        elif meal_type == 'lunch':
            lunch = description
        elif meal_type == 'dinner':
            dinner = description
        elif meal_type == 'snack':
            snacks.append(description)

    return breakfast, lunch, dinner, len(snacks), '; '.join(snacks)


def _extract_medicines_column(medicines_for_day: RowsData) -> tuple[int, str]:
    """Формирует многострочную колонку с лекарствами за день.

    Args:
        medicines_for_day: Список записей лекарств за день.

    Returns:
        tuple[int, str]: Количество приемов лекарства и текстовая колонка.
    """
    items: list[str] = []
    for medicine in medicines_for_day:
        name = medicine['name']
        dosage = (medicine.get('dosage') or '').strip()
        items.append(f'{name} ({dosage})' if dosage else name)
    return len(items), '\n'.join(items)


def _extract_stool_columns(stools_for_day: RowsData) -> tuple[int, str]:
    """Формирует колонки отчёта по стулу для одной даты.

    Args:
        stools_for_day: Список оценок стула за день.

    Returns:
        tuple[int, str]: Количество походов в туалет и расшифровка оценок.
    """
    stool_count = 0
    descriptions: list[str] = []
    for stool in stools_for_day:
        quality = int(stool['quality'])
        if quality != 0:
            stool_count += 1
        quality_text = BRISTOL.get(quality, 'неизвестно')
        descriptions.append(f'{quality} — {quality_text}')
    return stool_count, '\n'.join(descriptions)


def _build_report_rows(data: RowData) -> list[list[Any]]:
    """Строит строки DataFrame для Excel-отчёта.

    Args:
        data: Сырые данные пользователя из репозитория.

    Returns:
        list[list[Any]]: Нормализованные строки отчёта.
    """
    meals: RowsData = data['meals']
    medicines: RowsData = data['medicines']
    stools: RowsData = data['stools']
    feelings: RowsData = data['feelings']
    water: RowsData = data['water']
    sleeps: RowsData = data['sleeps']

    all_dates = sorted(
        {
            *[row['date'] for row in meals],
            *[row['date'] for row in medicines],
            *[row['date'] for row in stools],
            *[row['date'] for row in feelings],
            *[row['date'] for row in water],
            *[row['date'] for row in sleeps],
        }
    )

    meals_by_date = _group_rows_by_date(meals)
    medicines_by_date = _group_rows_by_date(medicines)
    stools_by_date = _group_rows_by_date(stools)
    feelings_by_date = _group_rows_by_date(feelings)
    water_by_date = {
        row['date']: int(row['glasses_count'])
        for row in water
    }
    sleep_by_date = {
        row['date']: row
        for row in sleeps
    }

    report_rows: list[list[Any]] = []
    for day in all_dates:
        breakfast, lunch, dinner, snacks_count, snacks_text = (
            _extract_meal_columns(
                meals_by_date.get(day, []),
            )
        )
        medicines_count, medicines_text = _extract_medicines_column(
            medicines_by_date.get(day, []),
        )
        stool_count, stool_text = _extract_stool_columns(
            stools_by_date.get(day, []),
        )
        feelings_text = '\n'.join(
            row['description'] for row in feelings_by_date.get(day, [])
        )

        sleep_row = sleep_by_date.get(day, {})
        wakeup_time = sleep_row.get('wakeup_time', '')
        bed_time = sleep_row.get('bed_time', '')
        sleep_quality = sleep_row.get('quality_description') or ''

        report_rows.append(
            [
                _to_display(day),
                breakfast,
                lunch,
                dinner,
                snacks_count,
                snacks_text,
                medicines_count,
                medicines_text,
                stool_count,
                stool_text,
                feelings_text,
                wakeup_time,
                bed_time,
                sleep_quality,
                water_by_date.get(day, 0),
            ]
        )
    return report_rows


def generate_user_report_xlsx(user_id: int) -> io.BytesIO:
    """Формирует XLSX-отчёт по всей истории пользователя.

    Args:
        user_id: Идентификатор пользователя Telegram.

    Returns:
        io.BytesIO: Буфер Excel-файла, готовый к отправке в Telegram.
    """
    rows = _build_report_rows(fetch_all_for_report(user_id))
    report_dataframe = pd.DataFrame(rows, columns=REPORT_COLUMNS)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_dataframe.to_excel(
            writer,
            index=False,
            sheet_name='Статистика',
        )
        worksheet = writer.sheets['Статистика']
        _apply_worksheet_style(
            worksheet,
            total_rows=len(report_dataframe.index) + 1,
            total_columns=len(report_dataframe.columns),
        )

    output.seek(0)
    return output
